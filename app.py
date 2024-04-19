from flask import Flask, jsonify, request
from datetime import datetime, timedelta
from openpyxl import load_workbook
import os

# "use": "@liudonghua123/now-flask"

app = Flask(__name__)

# Function to calculate due date based on LMP
def calculate_due_date(lmp_date):
    # Assuming 40 weeks gestation period
    due_date = lmp_date + timedelta(weeks=40)
    return due_date

# Function to calculate trimester based on current week
def calculate_trimester(current_week):
    if current_week <= 13:
        return "First Trimester"
    elif current_week <= 26:
        return "Second Trimester"
    else:
        return "Third Trimester"

# Function to calculate current pregnancy week
def calculate_current_week(lmp_date):
    today = datetime.today()
    # Calculate difference in days between today and LMP
    days_since_lmp = (today - lmp_date).days
    # Calculate weeks since LMP
    current_week = days_since_lmp // 7
    return current_week

# Function to calculate remaining weeks until due date
def calculate_remaining_weeks(lmp_date_str):
    lmp_date = datetime.strptime(lmp_date_str, "%Y-%m-%d")
    due_date = calculate_due_date(lmp_date)
    today = datetime.today()
    # Calculate difference in days between today and due date
    days_until_due_date = (due_date - today).days
    # Calculate remaining weeks
    remaining_weeks = max(0, days_until_due_date // 7)
    return remaining_weeks

# Function to calculate days left until due date
def calculate_days_left(lmp_date_str):
    lmp_date = datetime.strptime(lmp_date_str, "%Y-%m-%d")
    due_date = calculate_due_date(lmp_date)
    today = datetime.today()
    # Calculate difference in days between today and due date
    days_left = max(0, (due_date - today).days)
    return days_left

# API endpoints
@app.route('/duedate/<lmp_date>')
def get_due_date(lmp_date):
    lmp_date = datetime.strptime(lmp_date, "%Y-%m-%d")
    due_date = calculate_due_date(lmp_date)
    return jsonify({"due_date": due_date})

@app.route('/trimester/<lmp_date>')
def get_trimester(lmp_date):
    lmp_date = datetime.strptime(lmp_date, "%Y-%m-%d")
    current_week = calculate_current_week(lmp_date)
    trimester = calculate_trimester(current_week)
    return jsonify({"trimester": trimester})

@app.route('/mass_progress/<int:week>')
def get_mass_for_week(week):
    if week < 8:
        return jsonify({"message": "Week is too small"}), 400

    try:
        # Get the directory of the current Python script
        script_dir = os.path.dirname(__file__)

        # Construct the path to the Excel file
        excel_file = os.path.join(script_dir, 'data.xlsx')
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        
        # Assuming 'Week' and 'Mass' are column headers
        week_column = 'A'  # Update with the column letter for 'Week' column
        mass_column = 'C'  # Update with the column letter for 'Mass' column
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == week:
                mass = float(row[2]) if row[2] is not None else None
                return jsonify({"mass": mass})
                
        return jsonify({"message": "Mass not found for the specified week"}), 404
    except Exception as e:
        print(f"Error occurred: {e}")
        return jsonify({"message": "Internal Server Error"}), 500


@app.route('/length_progress/<int:week>')
def get_length_for_week(week):
    if week < 8:
        return jsonify({"message": "Week is too small"}), 400

    try:
         # Get the directory of the current Python script
        script_dir = os.path.dirname(__file__)

        # Construct the path to the Excel file
        excel_file = os.path.join(script_dir, 'data.xlsx')
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        
        # Assuming 'Week' and 'Mass' are column headers
        week_column = 'A'  # Update with the column letter for 'Week' column
        length_column = 'B'  # Update with the column letter for 'Length' column
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == week:
                length = float(row[1]) if row[1] is not None else None
                return jsonify({"length": length})
                
        return jsonify({"message": "Length not found for the specified week"}), 404
    except Exception as e:
        print(f"Error occurred: {e}")
        return jsonify({"message": "Internal Server Error"}), 500



@app.route('/current_week/<lmp_date>')
def get_current_week(lmp_date):
    lmp_date = datetime.strptime(lmp_date, "%Y-%m-%d")
    current_week = calculate_current_week(lmp_date)
    return jsonify({"current_week": current_week})

@app.route('/remaining_weeks/<string:lmp_date>')
def get_remaining_weeks(lmp_date):
    remaining_weeks = calculate_remaining_weeks(lmp_date)
    return jsonify({"remaining_weeks": remaining_weeks})

@app.route('/days_left/<string:lmp_date>')
def get_days_left(lmp_date):
    days_left = calculate_days_left(lmp_date)
    return jsonify({"days_left": days_left})



"""MENSTRUAL TRACKING
    ENDPOINTS AND FUNCTIONS START HERE"""

# Function to calculate the current menstrual cycle phase
def calculate_cycle_phase(last_period_date_str):
    last_period_date = datetime.strptime(last_period_date_str, "%Y-%m-%d")
    today = datetime.today()
    cycle_length = 28  # Default menstrual cycle length (can be adjusted)
    # Calculate days since the last period
    days_since_last_period = (today - last_period_date).days
    # Calculate the day within the menstrual cycle
    cycle_day = (days_since_last_period % cycle_length) + 1
    # Determine the cycle phase based on the cycle day
    if cycle_day <= 5:
        return "Menstrual"
    elif cycle_day <= 14:
        return "Follicular"
    elif cycle_day <= 21:
        return "Ovulatory"
    else:
        return "Luteal"

# Function to estimate the chance of pregnancy
def calculate_pregnancy_chance(menstrual_cycle_length, current_cycle_day):
    
    # Define the fertile window based on the menstrual cycle length
    ovulation_day = int(menstrual_cycle_length * 0.14)  # Approximation of ovulation day
    fertile_window_start = ovulation_day - 5
    fertile_window_end = ovulation_day + 4
    
    # Check if the current day is within the fertile window
    if fertile_window_start <= current_cycle_day <= fertile_window_end:
        # If within fertile window, higher chance of pregnancy
        return "High"  # Output "high" if within fertile window
    else:
        # If outside fertile window, lower chance of pregnancy
        return "Low"  # Output "low" if outside fertile window

# Function to calculate the date of the next period
def calculate_next_period_date(last_period_date, cycle_length):
    last_period_date = datetime.strptime(last_period_date, "%Y-%m-%d")
    next_period_date = last_period_date + timedelta(days=cycle_length)
    return next_period_date.strftime("%Y-%m-%d")

# API endpoints for menstrual tracking
@app.route('/calculate_cycle_phase/<last_period_date_str>')
def get_cycle_phase(last_period_date_str):
    cycle_phase = calculate_cycle_phase(last_period_date_str)
    return jsonify({'cycle_phase': cycle_phase})

@app.route('/calculate_pregnancy_chance/<int:cycle_length>/<int:current_day>')
def calculate_pregnancy(cycle_length, current_day):
    pregnancy_chance = calculate_pregnancy_chance(cycle_length, current_day)
    return jsonify({'pregnancy_chance': pregnancy_chance})

@app.route('/next_period_date/<last_period_date_str>/<int:cycle_length>')
def calculate_next_period(last_period_date_str, cycle_length):
    next_period_date = calculate_next_period_date(last_period_date_str, cycle_length)
    return jsonify({'next_period_date': next_period_date})

if __name__ == '__main__':
    app.run(debug=False)
